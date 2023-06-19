#!/usr/local/python3/bin/python3
# -*- coding:utf-8 -*-
import openpyxl

wb = openpyxl.load_workbook('/usr/local/webserver/DBA/zabbix_alert/monitor_server_list_20230405.xlsx')

#print(type(wb))

#ws = wb.active

#print(ws)

#print(wb.worksheets)
#print(wb.sheetnames)
#print(wb.get_sheet_by_name('Server List'))

##工作列表
sheet1 = wb['Server List']
print(sheet1.title)
print(sheet1.max_row)
print(sheet1.rows)
print(sheet1.columns)
print(sheet1.values)

print("===============")
cell = sheet1.cell(row=3,column=2)
print(cell)
print(cell.value)

print("**************")
data = sheet1.values
print(data)
print(list(data))

if __name__ == "__main__":
    pass
