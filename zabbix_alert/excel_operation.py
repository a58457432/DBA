#!/usr/local/python3/bin/python3
# -*- coding:utf-8 -*-

import openpyxl


excel_path = '/root/monitor_server_list_20230405.xlsx'
alert_relationship_sheet_name = 'Server List'
alert_first_header = '编号'
alert_headerlist = {
    'name':'Nost Name',
    'ip':'IP Address',
    'ping_wechat_users':'PING微信联系人',
    'ping_email_users':'PING邮件联系人',
    'ping_sms_users':'PING短信联系人',
    'performance_wechat_users': 'PERFORMANCE微信联系人',
    'performance_email_users': 'PERFORMANCE邮件联系人',
    'performance_sms_users': 'PERFORMANCE短信联系人',
    'application_wechat_users': 'APPLICATION微信联系人',
    'application_email_users': 'APPLICATION邮件联系人',
    'application_sms_users': 'APPLICATION短信联系人',
    'snmp_wechat_users': 'SNMP微信联系人',
    'snmp_email_users': 'SNMP邮件联系人',
    'snmp_sms_users': 'SNMP短信联系人',
}



class excel_operation:
    # excel操作类
    path = None
    sheet_name = None
    wb = None
    sheet = None
    max_row = None
    max_column = None
    # 标题和首列名称
    first_header = None
    headerlist = None

    def __init__(self,path,sheet_name):
        self.path = path
        self.sheet_name = sheet_name
        self.wb = openpyxl.load_workbook(path)
        # 读取成员列表
        self.sheet = self.wb.get_sheet_by_name(sheet_name)
        # 获取表的行和列总数
        self.max_row = self.sheet.max_row
        self.max_column = self.sheet.max_column


    # 获取标题所在行
    def get_data_start_row(self):
        for i in range(1,self.max_row + 1):
            if str(self.sheet.cell(row=i,column=1).value).strip() == self.first_header:
                data_start_row = i
        if data_start_row:return data_start_row

    # 获取所有需要获取的标题列名所在列
    def get_headers_column(self,data_start_row):
        header_column = {}
        # 把标题所在行的列循环
        for i in range(1,self.max_column + 1):
            column_value = str(self.sheet.cell(row=data_start_row,column=i).value).strip()
            # 把标题列名字典循环 判断列名 获得行号
            for key,value in self.headerlist.items():
                if column_value == value:
                    header_column[key] = i
        if header_column:return header_column

    # 获取指定行列的值
    def get_cell_value(self,row_num,column_num):
        value = self.sheet.cell(row=row_num,column=column_num).value
        if value:return str(value).strip()

    # 获取excel表内容 返回列表 列表为每行 里面嵌套字典 字典为列名和值
    def get_excel_data(self):
        # 获取标题所在行
        data_start_row = self.get_data_start_row()
        if data_start_row:
#            print('%s 表标题在第%s行' % (self.sheet_name, data_start_row))
            # 获取有用标签所在列
            headers_column = self.get_headers_column(data_start_row)
            if headers_column:
#                print(headers_column)
                # 循环每行 获得对应的列的值左操作（真实数据所在行是标题的下一行 所以+1,最大行数由于python原因需要+1）
                # 用户记录所有标签的列表（里面嵌套字典）
                datas_list = []
                for row_num in range(data_start_row + 1, self.max_row + 1):
                    row_dict = {}
                    # 只循环出需要数据所在的列
                    for header_key, column_num in headers_column.items():
                        cell_value = self.get_cell_value(row_num, column_num)
                        if cell_value: row_dict[header_key] = cell_value
                    # 把标签信息加入标签信息列表
                    if row_dict:
                        # 把标签信息加入标签信息列表
                        datas_list.append(row_dict)
        if datas_list: return datas_list

if __name__ == "__main__":
    alert_xl = excel_operation(excel_path, alert_relationship_sheet_name)
    alert_xl.first_header = alert_first_header
    alert_xl.headerlist = alert_headerlist
    

    alert_datas_list = alert_xl.get_excel_data()
    print(alert_datas_list[1])
#    print(len(alert_datas_list))    
    for k, v in alert_datas_list[1].items():
        print(k, '=' , v)
